from openpyxl import load_workbook
import requests
import json
from jvm_sever import JvmServer
import json_tools
import sys
import configparser

from file_logging import Logger
from config import ConfigOperation

class ExcelServer:
    def open_excel(self):
        wb = load_workbook(ConfigOperation.read_config().get("file", "path"))
        return wb

    def write_excel(self, wb, sheet, json_date, case_key=None, case_value=None):
        jvm_headers = JvmServer()
        case_index = 1

        row_index = 2
        if case_key != None:
            pass
        else:
            while row_index <= sheet.max_row:
                jvm_headers.open_JVM()
                Logger.logger().debug("第{case_index}次执行用例".format(case_index=case_index))
                # print("第{case_index}次执行用例".format(case_index=case_index))
                sys_name = sheet.cell(row=row_index, column=1).value                        # 系统名称
                case_id = sheet.cell(row=row_index, column=2).value
                # 用例ID
                interface_name = sheet.cell(row=row_index, column=3).value                  # 接口名称
                test_type = sheet.cell(row=row_index, column=4).value                       # 测试类型
                classify = sheet.cell(row=row_index, column=5).value                        # 分类
                case_name = sheet.cell(row=row_index, column=6).value                       # 用例名称
                case_leve = sheet.cell(row=row_index, column=7).value                       # 用例级别
                precondition = sheet.cell(row=row_index, column=8).value                    # 前置条件
                url = sheet.cell(row=row_index, column=9).value                             # 请求路径
                request_mode = sheet.cell(row=row_index, column=10).value                   # 请求方式(post or get)
                headers_text = sheet.cell(row=row_index, column=11).value                   # 请求头部
                request_body = sheet.cell(row=row_index, column=12).value                   # 请求body
                interface_expected_result = sheet.cell(row=row_index, column=13).value      # 期望返回结果
                data_expected_result = sheet.cell(row=row_index, column=14).value           # 数据库预期结果
                data_real_result = sheet.cell(row=row_index, column=16).value               # 数据库执行结果
                case_run_state = sheet.cell(row=row_index, column=17).value                 # 用例执行状态

                host = ConfigOperation.read_config().get("interface", "host")
                path = host + url
                headers = jvm_headers.get_data_api_header()

                # 将request_body转成json，判断endTime和beginTime是否存在，如存在，参数化赋值。
                data = json.loads(request_body)
                if "endTime" in data:
                    data["endTime"] = ConfigOperation.read_config().get("parameters", "endTime")
                if "beginTime" in data:
                    data["beginTime"] = ConfigOperation.read_config().get("parameters", "beginTime")
                body = str(data).replace("\'", "\"")
                # 此处data直接使用字符串，无需通过json.dumps()转换成Json。如转换，会报“JSON parse error: Cannot construct instance of `com.pagoda.disp.store.stats.rt.dto.ItemQueryParams”
                # 错误，分析可能是后台代码中已经处理了Json转换。所以后台需要接收一个str。如请求中转Json，那么后台就会参数类型错误。主观分析，未经求证。
                result = requests.post(path, data=body, headers=headers)

                if json_date == '接口返回结果':
                    sheet.cell(row=row_index, column=15).value = result.text                # 接口返回结果赋值，通过接口请求返回
                elif json_date == '期望返回结果':
                    sheet.cell(row=row_index, column=13).value = result.text
                wb.save(ConfigOperation.read_config().get("file", "path"))
                row_index = row_index + 1
                case_index = case_index + 1

        jvm_headers.close_JVM()
        # wb.save(r"testcase\用例all.xlsx")

    def get_case_dict(self, sheet, case_key=None, case_value=None):
        record_list = []
        row_index = 2

        if case_key != None:
            while row_index <= sheet.max_row:
                record_dict = {}
                column_index = 1
                while column_index <= sheet.max_column:
                    key_name = sheet.cell(row=1, column=column_index).value
                    value = sheet.cell(row=row_index, column=column_index).value
                    if key_name == "请求body" or key_name == "接口返回结果":
                        if value != None:
                            value = json.loads(value)
                    record_dict[key_name] = value
                    column_index = column_index + 1
                if case_key in record_dict.keys() and record_dict[case_key] == case_value:
                    record_list.append(record_dict)
                row_index = row_index + 1
        else:
            while row_index <= sheet.max_row:
                record_dict = {}
                column_index = 1
                while column_index <= sheet.max_column:
                    key_name = sheet.cell(row=1, column=column_index).value
                    value = sheet.cell(row=row_index, column=column_index).value
                    if key_name == "请求body" or key_name == "接口返回结果":
                        if value != None:
                            value = json.loads(value)
                    record_dict[key_name] = value
                    column_index = column_index + 1
                record_list.append(record_dict)
                row_index = row_index + 1

        return record_list

    # def get_case_tuple(self, sheet, case_name=None, case_value=None):
    #     record_list = []
    #     record_dict = self.get_case_dict(sheet, case_name, case_value)
    #     for record in record_dict:
    #         record_list_temp = [record["请求路径"], json.loads(record["请求body"]), json.loads(record["接口返回结果"])]
    #         record_list.append(record_list_temp)
    #     return record_list

    # def json_diff(self, wb, sheet):
    #     row_index = 2
    #     case_index = 1
    #     while row_index <= sheet.max_row:
    #         result_json = json.loads(sheet.cell(row=row_index, column=15).value)                        # 接口返回结果
    #         expect_json = json.loads(sheet.cell(row=row_index, column=13).value)                        # 期望返回结果
    #         result = json_tools.diff(result_json, expect_json)                                          # Json对比结果
    #         Logger.logger().debug("第{case_index}条用例Json对比结果：{result}".format(result=result, case_index=case_index))
    #         # print("第{case_index}条用例Json对比结果：{result}".format(result=result, case_index=case_index))
    #         if len(result) == 1 and result[0]["replace"] == "/timestamp":
    #             sheet.cell(row=row_index, column=17).value = "pass"
    #         else:
    #             sheet.cell(row=row_index, column=17).value = str(result)
    #         try:
    #             wb.save(ConfigOperation.read_config().get("file", "path"))
    #         except Exception as e:
    #             Logger.logger().error(e)
    #         row_index = row_index + 1
    #         case_index = case_index + 1

if __name__ == "__main__":
    excel_server = ExcelServer()
    wb = excel_server.open_excel()
    excel_server.write_excel(wb, wb[sys.argv[1]], sys.argv[2])


# 2、读取testlink测试用例